from openpyxl import Workbook


class SpreadSheet:
    def __init__(self, title: str = None):
        self.workbook = Workbook()
        self.workbook.active.title = title
        self.worksheet = self.workbook.active

    def add_dict(self, input_dict: dict, x: int = 0, y: int = 0, recursive_delimiter: str = "_"):
        """
        # add_dict
        Adds a table from a dictionary

        Args:
          input_dict: The dictionary that will be converted into a table
          x: The number of columns to the right that the top left of the table will be placed at. Must be an integer greater than or equal to zero.
          y: The number of rows down that the top left of the table will be placed at. Must be an integer greater than or equal to zero.
          recursive_delimiter: An optional parameter for handling recursive dictionaries.

        If input_dict contains a value that is not a str, int, float, bool, or dict, it will be converted to a str
        If input_dict contains a key that is not a str, it will be converted to a str
        """

        table_dict = self._format_dict(input_dict, recursive_delimiter)

        i = 0
        for k, v in table_dict.items():
            i += 1
            self.worksheet.cell(row=i+y, column=x, value=k)
            self.worksheet.cell(row=i+y, column=x+1, value=v)

    def _format_dict(self, input_dict: dict, recursive_delimiter: str = "_", prefix="") -> dict:
        table_dict = {}

        pre = prefix + recursive_delimiter if prefix != "" else prefix

        for k, v in input_dict.items():
            k = str(k)

            if isinstance(v, str):
                table_dict[pre + k] = v

            elif isinstance(v, bool):
                table_dict[pre + k] = "true" if v else "false"

            elif isinstance(v, int):
                table_dict[pre + k] = str(v)

            elif isinstance(v, float):
                table_dict[pre + k] = str(v)

            elif isinstance(v, dict):
                table_dict.update(self._format_dict(
                    v, recursive_delimiter, prefix=pre + k))

            else:
                table_dict[pre + k] = v.__str__()

        return table_dict

    def save_sheet(self, path: str = None):
        """
        Saves the current sheet to the path provided
        """
        self.workbook.save(path)
